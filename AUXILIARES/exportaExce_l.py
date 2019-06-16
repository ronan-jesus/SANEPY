# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import wx
from PIL import Image
import pandas as pd


        
        
class ExportaEstruturaExcel(object):
    """docstring for ExportaEstruturaExcel"""
    def __init__(self, model, estrutura, event):
        super(ExportaEstruturaExcel, self).__init__()
        self.model = model
        self.Estrutura = estrutura
        self.event = event
        self.workbook = Workbook()
        self.caminho = None
        self.SalvaEstrutura()
        
    
    
    
    def EscreveDadosBarras(self):
        print "Entou aqui 2"        
        
        for b in self.Estrutura.LISTA_BARRAS:           
            
            worksheet = self.workbook.create_sheet(title = u'Barra '+str(b.numero))
            worksheet = self.workbook.active
            
           
            worksheet['A1']= u'COORDENADAS DOS NÓS'
            worksheet.column_dimensions["A"].width = 60
            worksheet.merge_cells('A1:D1')
            worksheet['A1']= u'COORDENADAS DOS NÓS'
#            worksheet.write_merge(1,1,0,1, u'NÓS')
#            worksheet.write(1,2, u'PosX')
#            worksheet.write(1,3, u'PosY')
#            worksheet.write(1,4, u'PosZ')
#
#            worksheet.write(2,0, u'Inicial')
#            worksheet.write(2,1, str(b.NO1.numero))
#            worksheet.write(3,0, u'Final')
#            worksheet.write(3,1, str(b.NO2.numero))
#
#            worksheet.write(2,2, str(b.NO1.GetPosNo()[0]).replace(".",","))
#            worksheet.write(2,3, str(b.NO1.GetPosNo()[1]).replace(".",","))
#            worksheet.write(2,4, str(b.NO1.GetPosNo()[2]).replace(".",","))
#
#            worksheet.write(3,2, str(b.NO2.GetPosNo()[0]).replace(".",","))
#            worksheet.write(3,3, str(b.NO2.GetPosNo()[1]).replace(".",","))
#            worksheet.write(3,4, str(b.NO2.GetPosNo()[2]).replace(".",","))
#
#
#            ####### INICIO DESCRICAO DAS CARACTERISTICAS DA BARRA #################
#            worksheet.write_merge(0,0,7,8, u'GRANDEZAS', xlwt.easyxf("align: horiz center"))
#            worksheet.write(0,9, u'VALOR', xlwt.easyxf("align: horiz center"))
#            worksheet.write(0,10, u'UNIDADE', xlwt.easyxf("align: horiz center"))
#
#            worksheet.write(1,7, u'Compr. Horizon-X')
#            worksheet.write(2,7, u'Compr. Horizon-Y')
#            worksheet.write(3,7, u'Compr. Vertica-Z')
#            worksheet.write(4,7, u'Compr. Total')            
#            worksheet.write(1,8, u'Lx')
#            worksheet.write(2,8, u'Ly')
#            worksheet.write(3,8, u'Lz')
#            worksheet.write(4,8, u'L')
#
#            worksheet.write(1,9, str(b.GetComprimentos()[0]).replace(".",","))
#            worksheet.write(2,9, str(b.GetComprimentos()[1]).replace(".",","))
#            worksheet.write(3,9, str(b.GetComprimentos()[2]).replace(".",","))
#            worksheet.write(4,9, str(b.Get_L()).replace(".",","))
#
#            worksheet.write(1,10, str(self.model.TextoUnidadeComprimento))
#            worksheet.write(2,10, str(self.model.TextoUnidadeComprimento))
#            worksheet.write(3,10, str(self.model.TextoUnidadeComprimento))
#            worksheet.write(4,10, str(self.model.TextoUnidadeComprimento))
#
#            worksheet.write_merge(6,6,0,4, u'CARACTERISTICAS DA BARRA', xlwt.easyxf("align: horiz center"))
#            worksheet.write_merge(7,7,0,4, u'Propriedades do Material', xlwt.easyxf("align: horiz center"))
#            worksheet.write(8,0, u'E')
#            worksheet.write(8,1, u'=')
#            worksheet.write_merge(8,8,2,3, str(b.GetE()).replace(".",","))
#            worksheet.write(8,4, str(self.model.TextoUnidadeForca+'/'+self.model.TextoUnidadeComprimento+'^2'))
#            worksheet.write(9,0, u'G')
#            worksheet.write(9,1, u'=')
#            worksheet.write_merge(9,9,2,3, str(b.GetG()).replace(".",","))
#            worksheet.write(9,4, str(self.model.TextoUnidadeForca+'/'+self.model.TextoUnidadeComprimento+'^2'))
#
#
#
#            worksheet.write_merge(11,11,0,4, u'Propriedades da Seção', xlwt.easyxf("align: horiz center"))
#            worksheet.write(12,0, u'A')
#            worksheet.write(12,1, u'=')
#            worksheet.write_merge(12,12,2,3, str(b.GetA()).replace(".",","))
#            worksheet.write(12,4, str(self.model.TextoUnidadeComprimento+'^2'))
#            worksheet.write(13,0, u'Ix')
#            worksheet.write(13,1, u'=')
#            worksheet.write_merge(13,13,2,3, str(b.GetIx()).replace(".",","))
#            worksheet.write(13,4, str(self.model.TextoUnidadeComprimento+'^4'))
#            worksheet.write(14,0, u'Iy')
#            worksheet.write(14,1, u'=')
#            worksheet.write_merge(14,14,2,3, str(b.GetIy()).replace(".",","))
#            worksheet.write(14,4, str(self.model.TextoUnidadeComprimento+'^4'))
#            worksheet.write(15,0, u'Iz')
#            worksheet.write(15,1, u'=')
#            worksheet.write_merge(15,15,2,3, str(b.GetIz()).replace(".",","))
#            worksheet.write(15,4, str(self.model.TextoUnidadeComprimento+'^4'))
#            ####### FIM DESCRICAO DAS CARACTERISTICAS DA BARRA #################
#
#
#            ####### INICIO DESCRICAO DA MATRIZ DE RIGIDEZ DA BARRA #################
#            worksheet.write_merge(19,19,9,20, u'MATRIZ DE RIGIDEZ NO SISTEMA LOCAL', xlwt.easyxf("align: horiz center"))
#            worksheet.write(20,9, u'Uxi')
#            worksheet.write(20,10, u'Uyi')
#            worksheet.write(20,11, u'Uzi')
#            worksheet.write(20,12, u'θxi')
#            worksheet.write(20,13, u'θyi')
#            worksheet.write(20,14, u'θzi')
#            worksheet.write(20,15, u'Uxj')
#            worksheet.write(20,16, u'Uyj')
#            worksheet.write(20,17, u'Uzj')
#            worksheet.write(20,18, u'θxj')
#            worksheet.write(20,19, u'θyj')
#            worksheet.write(20,20, u'θzj')
#
#            worksheet.write(21,8, u'Uxi')
#            worksheet.write(22,8, u'Uyi')
#            worksheet.write(23,8, u'Uzi')
#            worksheet.write(24,8, u'θxi')
#            worksheet.write(25,8, u'θyi')
#            worksheet.write(26,8, u'θzi')
#            worksheet.write(27,8, u'Uxj')
#            worksheet.write(28,8, u'Uyj')
#            worksheet.write(29,8, u'Uzj')
#            worksheet.write(30,8, u'θxj')
#            worksheet.write(31,8, u'θyj')
#            worksheet.write(32,8, u'θzj')
#
#            worksheet.write(27,7, u'K_local ==>')
#            matrizRigidezLocal = b.GetMatrizRigidezLocal()
#            for i in range(12):
#                for j in range(12):
#                    worksheet.write(21+i,9+j, str(matrizRigidezLocal[i][j]).replace(".",","))
#            ####### FIM DESCRICAO DA MATRIZ DE RIGIDEZ DA BARRA #################
#
#
#            ####### INICIO DESCRICAO DA MATRIZ DE ROTACAO DA BARRA #################
#            worksheet.write_merge(35,35,9,20, u'MATRIZ DE ROTACAO', xlwt.easyxf("align: horiz center"))
#
#            worksheet.write(42,8, u'R ==>')
#            for i in range(12):
#                for j in range(12):
#                    worksheet.write(36+i,9+j, str(b.R[i][j]).replace(".",","))
#            ####### FIM DESCRICAO DA MATRIZ DE ROTACAO DA BARRA #################
#                    
#            ####### INICIO DESCRICAO DA MATRIZ DE ROTACAO TRANSPOSTA DA BARRA #################
#            worksheet.write_merge(50,50,9,20, u'MATRIZ DE ROTACAO TRANSPOSTA', xlwt.easyxf("align: horiz center"))
#
#            worksheet.write(57,8, u'RT ==>')
#            for i in range(12):
#                for j in range(12):
#                    worksheet.write(51+i,9+j, str(b.RT[i][j]).replace(".",","))
#            ####### FIM DESCRICAO DA MATRIZ DE ROTACAO TRANSPOSTA DA BARRA #################
#                    
#                    
#            ####### INICIO DESCRICAO DA MATRIZ DE RIGIDEZ DA BARRA #################
#            worksheet.write_merge(65,65,9,20, u'MATRIZ DE RIGIDEZ NO SISTEMA GLOBAL', xlwt.easyxf("align: horiz center"))
#            worksheet.write(66,9, u'Uxi')
#            worksheet.write(66,10, u'Uyi')
#            worksheet.write(66,11, u'Uzi')
#            worksheet.write(66,12, u'θxi')
#            worksheet.write(66,13, u'θyi')
#            worksheet.write(66,14, u'θzi')
#            worksheet.write(66,15, u'Uxj')
#            worksheet.write(66,16, u'Uyj')
#            worksheet.write(66,17, u'Uzj')
#            worksheet.write(66,18, u'θxj')
#            worksheet.write(66,19, u'θyj')
#            worksheet.write(66,20, u'θzj')
#
#            worksheet.write(67,8, u'Uxi')
#            worksheet.write(68,8, u'Uyi')
#            worksheet.write(69,8, u'Uzi')
#            worksheet.write(70,8, u'θxi')
#            worksheet.write(71,8, u'θyi')
#            worksheet.write(72,8, u'θzi')
#            worksheet.write(73,8, u'Uxj')
#            worksheet.write(74,8, u'Uyj')
#            worksheet.write(75,8, u'Uzj')
#            worksheet.write(76,8, u'θxj')
#            worksheet.write(77,8, u'θyj')
#            worksheet.write(78,8, u'θzj')
#
#            worksheet.write(73,7, u'K_Global ==>')
#            matrizRigidezGlobal = b.GetMatrizRigidezGlobal()
#            for i in range(12):
#                for j in range(12):
#                    worksheet.write(67+i,9+j, str(matrizRigidezGlobal[i][j]).replace(".",","))
#            ####### FIM DESCRICAO DA MATRIZ DE RIGIDEZ DA BARRA #################
#            
#            ####### INICIO DESCRICAO DAS FORCAS DE ENGASTAMENTO PERFEITO LOCAL DA BARRA #################
#            worksheet.write_merge(80,80,0,8, u'VETOR DE CARGAS DE ENGASTAMENTO PERFEITO - SISTEMA LOCAL ')
#            worksheet.write(88,4, u'fengL ==>')
#            worksheet.write(83,5, u"Fxi =")
#            worksheet.write(84,5, u"Fyi =")
#            worksheet.write(85,5, u"Fzi =")
#            worksheet.write(86,5, u"Mxi =")
#            worksheet.write(87,5, u"Myi =")
#            worksheet.write(88,5, u"Mzi =")
#            worksheet.write(89,5, u"Fxj =")
#            worksheet.write(90,5, u"Fyj =")
#            worksheet.write(91,5, u"Fzj =")
#            worksheet.write(92,5, u"Mxj =")
#            worksheet.write(93,5, u"Myj =")
#            worksheet.write(94,5, u"Mzj =")
#            
#            worksheet.write(83,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(84,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(85,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(86,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            worksheet.write(87,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            worksheet.write(88,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            worksheet.write(89,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(90,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(91,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(92,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            worksheet.write(93,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            worksheet.write(94,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            fengLocal = b.GetFengLocal()
#            for i in range(12):
#                worksheet.write(83+i,6, str(fengLocal[i]).replace(".",","))
#            
#            #Insere imagem da barra com a carga distribuida
#            worksheet.insert_bitmap(r"img/barraCargaDist.bmp", 84, 0)
#            
#            worksheet.write(90,0, u"qy")
#            worksheet.write(90,1, u"=")
#            worksheet.write(91,0, u"qz")
#            worksheet.write(91,1, u"=")
#            worksheet.write(92,0, u"L")
#            worksheet.write(92,1, u"=")
#            worksheet.write(90,2, str(b.GetQy()))
#            worksheet.write(91,2, str(b.GetQz()))
#            worksheet.write(92,2, str(b.Get_L()))
#            worksheet.write(90,3, str(self.model.TextoUnidadeForca))
#            worksheet.write(91,3, str(self.model.TextoUnidadeForca))
#            worksheet.write(92,3, str(self.model.TextoUnidadeComprimento))
#            ####### FIM DESCRICAO DAS FORCAS DE ENGASTAMENTO PERFEITO LOCAL DA BARRA #################
#            
#            
#            ####### INICIO DESCRICAO DAS FORCAS DE ENGASTAMENTO PERFEITO GLABAL DA BARRA #################
#            worksheet.write_merge(96,96,0,8, u'VETOR DE CARGAS DE ENGASTAMENTO PERFEITO - SISTEMA GLOBAL ')
#            worksheet.write(103,4, u'fengG ==>')
#            worksheet.write(98,5, u"Fxi =")
#            worksheet.write(99,5, u"Fyi =")
#            worksheet.write(100,5, u"Fzi =")
#            worksheet.write(101,5, u"Mxi =")
#            worksheet.write(102,5, u"Myi =")
#            worksheet.write(103,5, u"Mzi =")
#            worksheet.write(104,5, u"Fxj =")
#            worksheet.write(105,5, u"Fyj =")
#            worksheet.write(106,5, u"Fzj =")
#            worksheet.write(107,5, u"Mxj =")
#            worksheet.write(108,5, u"Myj =")
#            worksheet.write(109,5, u"Mzj =")
#            
#            worksheet.write(98,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(99,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(100,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(101,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            worksheet.write(102,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            worksheet.write(103,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            worksheet.write(104,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(105,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(106,7, str(self.model.TextoUnidadeForca))
#            worksheet.write(107,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            worksheet.write(108,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            worksheet.write(109,7, str(self.model.TextoUnidadeForca+'.'+self.model.TextoUnidadeComprimento))
#            fengGlobal = b.GetFengGlobal()
#            for i in range(12):
#                worksheet.write(98+i,6, str(fengGlobal[i]).replace(".",","))
#            ####### FIM DESCRICAO DAS FORCAS DE ENGASTAMENTO PERFEITO GLOBAL DA BARRA #################
            
        
            
    def SalvaMatrizGlobal(self, caminho):
        RIG_GLOBAL = self.Estrutura.MATRIZ_RIG_GLOBAL
        
        book = load_workbook(caminho)
        writer = pd.ExcelWriter(caminho, engine = 'openpyxl')
        writer.book = book
        
        ## convert your array into a dataframe
        df = pd.DataFrame (RIG_GLOBAL)
        

        df.to_excel(writer, sheet_name = u'Sheet')   
               
    def SalvaEstrutura(self):    
        """
        Funcao responsavel por abrir a janela de salvamento, escolhe o nome e local
        do arquivo, apos faz o salvamento.
        """
        dlg = wx.FileDialog(self.event.GetEventObject().GetParent(), "Save project as...", os.getcwd(), "", "*.xlsx", wx.SAVE|wx.FD_OVERWRITE_PROMPT)
        result = dlg.ShowModal()
        inFile = dlg.GetPath()
        dlg.Destroy()

        if result == wx.ID_OK:  #Botao salvar foi pressionado            
            # Abra o arquivo (leitura)
            try:
                print "Entou aqui 1"
                self.EscreveDadosBarras()
                self.workbook.save(inFile)                
                #self.SalvaMatrizGlobal(inFile)
            except Exception as e:                
                print e                
            finally:
                pass
            return True

        elif result == wx.ID_CANCEL:    #Qualque um dos botoes cancelar ou fechar janela
            return False
        